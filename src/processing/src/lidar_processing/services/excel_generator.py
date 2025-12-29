"""
Excel Report Generation Service.

This module provides Excel workbook generation for forest inventory
using openpyxl. Creates multi-sheet workbooks with formatted tables,
charts, and summary statistics.
"""

from __future__ import annotations

import io
import logging
from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING, Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

    from lidar_processing.models import (
        InventorySummary,
        ProjectInfo,
        ReportOptions,
        SpeciesMetrics,
        StandMetrics,
        TreeMetrics,
    )

logger = logging.getLogger(__name__)

# Color definitions
HEADER_FILL = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
ACCENT_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(color="2E7D32", bold=True, size=14)
METRIC_FONT = Font(color="2E7D32", bold=True, size=18)
NORMAL_FONT = Font(color="333333", size=10)

# Border styles
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
HEADER_BORDER = Border(
    left=Side(style="thin", color="2E7D32"),
    right=Side(style="thin", color="2E7D32"),
    top=Side(style="thin", color="2E7D32"),
    bottom=Side(style="medium", color="2E7D32"),
)


class ExcelGenerator:
    """
    Generates Excel workbooks for forest inventory reports.

    Creates multi-sheet workbooks with:
    - Summary sheet with key metrics
    - Tree inventory with all measurements
    - Species summary with aggregated statistics
    - Stand summary (if applicable)
    - Embedded charts
    """

    def __init__(self) -> None:
        """Initialize the Excel generator."""
        pass

    def generate_workbook(
        self,
        output_path: str | Path,
        project_info: ProjectInfo,
        summary: InventorySummary,
        trees: list[TreeMetrics],
        options: ReportOptions,
        species_metrics: list[SpeciesMetrics] | None = None,
        stand_metrics: list[StandMetrics] | None = None,
        charts: dict[str, bytes] | None = None,
    ) -> str:
        """
        Generate a complete Excel workbook.

        Args:
            output_path: Path for the output Excel file.
            project_info: Project information.
            summary: Inventory summary statistics.
            trees: List of tree metrics.
            options: Report generation options.
            species_metrics: Optional species composition data.
            stand_metrics: Optional stand-level metrics.
            charts: Dictionary of chart names to PNG bytes.

        Returns:
            Path to the generated Excel file.
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()

        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)

        # Create sheets
        summary_ws = wb.create_sheet("Summary")
        self._create_summary_sheet(
            summary_ws, project_info, summary, options, charts
        )

        if options.include_tree_list:
            tree_ws = wb.create_sheet("Tree Inventory")
            self._create_tree_inventory_sheet(tree_ws, trees, options)

        if options.include_species_summary and species_metrics:
            species_ws = wb.create_sheet("Species Summary")
            self._create_species_sheet(species_ws, species_metrics)

        if options.include_stand_summary and stand_metrics:
            stand_ws = wb.create_sheet("Stand Summary")
            self._create_stand_sheet(stand_ws, stand_metrics)

        # Save workbook
        wb.save(str(output_path))
        logger.info("Generated Excel workbook: %s", output_path)

        return str(output_path)

    def _create_summary_sheet(
        self,
        ws: Worksheet,
        project_info: ProjectInfo,
        summary: InventorySummary,
        options: ReportOptions,
        charts: dict[str, bytes] | None,
    ) -> None:
        """Create the summary sheet with key metrics and charts."""
        # Title
        ws.merge_cells("A1:F1")
        title_cell = ws["A1"]
        title_cell.value = "Forest Inventory Report"
        title_cell.font = Font(color="2E7D32", bold=True, size=20)
        title_cell.alignment = Alignment(horizontal="center")

        # Project info section
        ws.merge_cells("A3:F3")
        ws["A3"].value = project_info.project_name
        ws["A3"].font = TITLE_FONT
        ws["A3"].alignment = Alignment(horizontal="center")

        row = 5
        project_data = [
            ("Client", project_info.client_name),
            ("Location", project_info.location),
            ("Project ID", project_info.project_id),
            ("Survey Date", project_info.survey_date.strftime("%Y-%m-%d") if project_info.survey_date else None),
            ("Analysis Date", project_info.analysis_date.strftime("%Y-%m-%d")),
            ("Coordinate System", project_info.coordinate_system),
        ]

        for label, value in project_data:
            if value:
                ws[f"A{row}"].value = label
                ws[f"A{row}"].font = Font(bold=True)
                ws[f"B{row}"].value = value
                row += 1

        row += 1

        # Key metrics section
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"].value = "Key Metrics"
        ws[f"A{row}"].font = TITLE_FONT
        row += 2

        # Metrics in a grid
        metrics = [
            ("Total Trees", f"{summary.total_trees:,}"),
            ("Survey Area", f"{summary.total_area_hectares:.2f} ha"),
            ("Stems per Hectare", f"{summary.stems_per_hectare:.0f}"),
            ("Mean Height", f"{summary.mean_height:.2f} m"),
            ("Max Height", f"{summary.max_height:.2f} m"),
            ("Height Std Dev", f"{summary.std_height:.2f} m"),
        ]

        if summary.mean_dbh:
            metrics.append(("Mean DBH", f"{summary.mean_dbh:.1f} cm"))
        if summary.basal_area_per_hectare:
            metrics.append(("Basal Area", f"{summary.basal_area_per_hectare:.2f} m2/ha"))
        if summary.total_biomass:
            metrics.append(("Total Biomass", f"{summary.total_biomass / 1000:.2f} t"))
        if summary.total_carbon:
            metrics.append(("Carbon Stock", f"{summary.total_carbon / 1000:.2f} t C"))
        if summary.co2_equivalent:
            metrics.append(("CO2 Equivalent", f"{summary.co2_equivalent / 1000:.2f} t CO2"))
        if summary.species_count:
            metrics.append(("Species Count", str(summary.species_count)))

        # Display metrics in 2 columns
        col_offset = 0
        start_row = row
        for i, (label, value) in enumerate(metrics):
            if i > 0 and i % 6 == 0:
                col_offset = 3
                row = start_row

            col_a = get_column_letter(1 + col_offset)
            col_b = get_column_letter(2 + col_offset)

            ws[f"{col_a}{row}"].value = label
            ws[f"{col_a}{row}"].font = Font(bold=True)
            ws[f"{col_a}{row}"].fill = ACCENT_FILL
            ws[f"{col_a}{row}"].border = THIN_BORDER

            ws[f"{col_b}{row}"].value = value
            ws[f"{col_b}{row}"].font = METRIC_FONT
            ws[f"{col_b}{row}"].alignment = Alignment(horizontal="right")
            ws[f"{col_b}{row}"].border = THIN_BORDER

            row += 1

        row = max(row, start_row + 6) + 2

        # Add charts as images
        if charts and options.include_charts:
            chart_row = row

            if "species_pie" in charts:
                try:
                    img = XLImage(io.BytesIO(charts["species_pie"]))
                    img.width = 400
                    img.height = 300
                    ws.add_image(img, f"A{chart_row}")
                    chart_row += 18
                except Exception as e:
                    logger.warning("Failed to add species chart: %s", e)

            if "height_histogram" in charts:
                try:
                    img = XLImage(io.BytesIO(charts["height_histogram"]))
                    img.width = 400
                    img.height = 300
                    ws.add_image(img, f"A{chart_row}")
                except Exception as e:
                    logger.warning("Failed to add height chart: %s", e)

        # Set column widths
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 5
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 15

    def _create_tree_inventory_sheet(
        self,
        ws: Worksheet,
        trees: list[TreeMetrics],
        options: ReportOptions,
    ) -> None:
        """Create the tree inventory sheet with all tree data."""
        # Headers
        headers = [
            "Tree ID",
            "X Coordinate",
            "Y Coordinate",
            "Height (m)",
            "Crown Diameter (m)",
            "Crown Area (m2)",
            "Crown Base Height (m)",
            "DBH (cm)",
            "Biomass (kg)",
            "Point Count",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = HEADER_BORDER
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_num, tree in enumerate(trees, 2):
            ws.cell(row=row_num, column=1).value = tree.tree_id
            ws.cell(row=row_num, column=2).value = round(tree.x, 3)
            ws.cell(row=row_num, column=3).value = round(tree.y, 3)
            ws.cell(row=row_num, column=4).value = round(tree.height, 2)
            ws.cell(row=row_num, column=5).value = (
                round(tree.crown_diameter, 2) if tree.crown_diameter else None
            )
            ws.cell(row=row_num, column=6).value = (
                round(tree.crown_area, 2) if tree.crown_area else None
            )
            ws.cell(row=row_num, column=7).value = (
                round(tree.crown_base_height, 2) if tree.crown_base_height else None
            )
            ws.cell(row=row_num, column=8).value = (
                round(tree.dbh_estimated, 1) if tree.dbh_estimated else None
            )
            ws.cell(row=row_num, column=9).value = (
                round(tree.biomass_estimated, 0) if tree.biomass_estimated else None
            )
            ws.cell(row=row_num, column=10).value = tree.point_count

            # Apply alternating row colors
            if row_num % 2 == 0:
                for col in range(1, 11):
                    ws.cell(row=row_num, column=col).fill = ACCENT_FILL

            # Apply borders
            for col in range(1, 11):
                ws.cell(row=row_num, column=col).border = THIN_BORDER
                ws.cell(row=row_num, column=col).font = NORMAL_FONT

        # Auto-fit columns
        self._auto_fit_columns(ws, headers)

        # Freeze header row
        ws.freeze_panes = "A2"

    def _create_species_sheet(
        self,
        ws: Worksheet,
        species_metrics: list[SpeciesMetrics],
    ) -> None:
        """Create the species summary sheet."""
        # Headers
        headers = [
            "Species",
            "Tree Count",
            "Percentage (%)",
            "Mean Height (m)",
            "Mean DBH (cm)",
            "Mean Crown (m)",
            "Total Basal Area (m2)",
            "Total Biomass (kg)",
            "Total Carbon (kg)",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = HEADER_BORDER
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_num, species in enumerate(species_metrics, 2):
            ws.cell(row=row_num, column=1).value = species.species_name
            ws.cell(row=row_num, column=2).value = species.tree_count
            ws.cell(row=row_num, column=3).value = round(species.percentage, 2)
            ws.cell(row=row_num, column=4).value = round(species.mean_height, 2)
            ws.cell(row=row_num, column=5).value = (
                round(species.mean_dbh, 1) if species.mean_dbh else None
            )
            ws.cell(row=row_num, column=6).value = (
                round(species.mean_crown_diameter, 2) if species.mean_crown_diameter else None
            )
            ws.cell(row=row_num, column=7).value = (
                round(species.total_basal_area, 4) if species.total_basal_area else None
            )
            ws.cell(row=row_num, column=8).value = (
                round(species.total_biomass, 0) if species.total_biomass else None
            )
            ws.cell(row=row_num, column=9).value = (
                round(species.total_carbon, 0) if species.total_carbon else None
            )

            # Apply alternating row colors
            if row_num % 2 == 0:
                for col in range(1, 10):
                    ws.cell(row=row_num, column=col).fill = ACCENT_FILL

            # Apply borders
            for col in range(1, 10):
                ws.cell(row=row_num, column=col).border = THIN_BORDER
                ws.cell(row=row_num, column=col).font = NORMAL_FONT

        # Add totals row
        total_row = len(species_metrics) + 2
        ws.cell(row=total_row, column=1).value = "TOTAL"
        ws.cell(row=total_row, column=1).font = Font(bold=True)
        ws.cell(row=total_row, column=2).value = sum(s.tree_count for s in species_metrics)
        ws.cell(row=total_row, column=2).font = Font(bold=True)
        ws.cell(row=total_row, column=3).value = 100.0
        ws.cell(row=total_row, column=3).font = Font(bold=True)

        total_biomass = sum(s.total_biomass or 0 for s in species_metrics)
        total_carbon = sum(s.total_carbon or 0 for s in species_metrics)
        ws.cell(row=total_row, column=8).value = round(total_biomass, 0)
        ws.cell(row=total_row, column=8).font = Font(bold=True)
        ws.cell(row=total_row, column=9).value = round(total_carbon, 0)
        ws.cell(row=total_row, column=9).font = Font(bold=True)

        for col in range(1, 10):
            ws.cell(row=total_row, column=col).border = HEADER_BORDER

        # Auto-fit columns
        self._auto_fit_columns(ws, headers)

        # Add pie chart for species distribution
        if len(species_metrics) > 0:
            chart = PieChart()
            chart.title = "Species Distribution"

            data_end = min(len(species_metrics) + 1, 11)  # Max 10 species in chart
            data = Reference(ws, min_col=2, min_row=1, max_row=data_end)
            labels = Reference(ws, min_col=1, min_row=2, max_row=data_end)

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(labels)
            chart.width = 15
            chart.height = 10

            ws.add_chart(chart, f"K2")

        # Freeze header row
        ws.freeze_panes = "A2"

    def _create_stand_sheet(
        self,
        ws: Worksheet,
        stand_metrics: list[StandMetrics],
    ) -> None:
        """Create the stand summary sheet."""
        # Headers
        headers = [
            "Stand ID",
            "Stand Name",
            "Area (ha)",
            "Tree Count",
            "Stems/ha",
            "Basal Area (m2/ha)",
            "Mean Height (m)",
            "Dominant Height (m)",
            "Mean DBH (cm)",
            "QMD (cm)",
            "Volume (m3/ha)",
            "Biomass (kg/ha)",
            "Carbon (kg/ha)",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = HEADER_BORDER
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_num, stand in enumerate(stand_metrics, 2):
            ws.cell(row=row_num, column=1).value = stand.stand_id
            ws.cell(row=row_num, column=2).value = stand.stand_name
            ws.cell(row=row_num, column=3).value = round(stand.area_hectares, 2)
            ws.cell(row=row_num, column=4).value = stand.tree_count
            ws.cell(row=row_num, column=5).value = round(stand.stems_per_hectare, 0)
            ws.cell(row=row_num, column=6).value = round(stand.basal_area_per_hectare, 2)
            ws.cell(row=row_num, column=7).value = round(stand.mean_height, 2)
            ws.cell(row=row_num, column=8).value = (
                round(stand.dominant_height, 2) if stand.dominant_height else None
            )
            ws.cell(row=row_num, column=9).value = (
                round(stand.mean_dbh, 1) if stand.mean_dbh else None
            )
            ws.cell(row=row_num, column=10).value = (
                round(stand.quadratic_mean_dbh, 1) if stand.quadratic_mean_dbh else None
            )
            ws.cell(row=row_num, column=11).value = (
                round(stand.volume_per_hectare, 1) if stand.volume_per_hectare else None
            )
            ws.cell(row=row_num, column=12).value = (
                round(stand.biomass_per_hectare, 0) if stand.biomass_per_hectare else None
            )
            ws.cell(row=row_num, column=13).value = (
                round(stand.carbon_per_hectare, 0) if stand.carbon_per_hectare else None
            )

            # Apply alternating row colors
            if row_num % 2 == 0:
                for col in range(1, 14):
                    ws.cell(row=row_num, column=col).fill = ACCENT_FILL

            # Apply borders
            for col in range(1, 14):
                ws.cell(row=row_num, column=col).border = THIN_BORDER
                ws.cell(row=row_num, column=col).font = NORMAL_FONT

        # Add bar chart for stand comparison
        if len(stand_metrics) > 1:
            chart = BarChart()
            chart.title = "Stand Comparison - Stems per Hectare"
            chart.type = "col"
            chart.style = 10

            data = Reference(ws, min_col=5, min_row=1, max_row=len(stand_metrics) + 1)
            categories = Reference(ws, min_col=1, min_row=2, max_row=len(stand_metrics) + 1)

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            chart.shape = 4
            chart.width = 15
            chart.height = 10

            ws.add_chart(chart, "O2")

        # Auto-fit columns
        self._auto_fit_columns(ws, headers)

        # Freeze header row
        ws.freeze_panes = "A2"

    def _auto_fit_columns(
        self,
        ws: Worksheet,
        headers: list[str],
    ) -> None:
        """
        Auto-fit column widths based on content.

        Args:
            ws: Worksheet to adjust.
            headers: Header row for minimum widths.
        """
        for col_num, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)

            # Start with header width
            max_length = len(str(header))

            # Check data in column (sample first 100 rows)
            for row in range(2, min(ws.max_row + 1, 102)):
                cell_value = ws.cell(row=row, column=col_num).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))

            # Set width with padding
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width
