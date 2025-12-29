"""
Carbon Report Generator Service.

Sprint 25-30: Carbon Stock Estimation

Generates VCS, CAR, and ACR protocol-compliant carbon stock reports
in PDF and Excel formats.
"""

from __future__ import annotations

import io
import logging
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

from lidar_processing.services.carbon_stock_estimator import (
    CarbonPool,
    CarbonProtocol,
    PoolType,
    ProjectCarbonStock,
    TreeCarbonEstimate,
    UncertaintyEstimate,
)

logger = logging.getLogger(__name__)


@dataclass
class CarbonReportConfig:
    """Configuration for carbon report generation."""
    include_methodology: bool = True
    include_uncertainty: bool = True
    include_audit_trail: bool = True
    include_tree_level: bool = False  # Can be large
    include_charts: bool = True
    include_credits: bool = True
    unit_system: str = "metric"  # 'metric' or 'imperial'


@dataclass
class CarbonReport:
    """Generated carbon report."""
    report_id: str
    project_id: str
    analysis_id: str
    protocol: CarbonProtocol

    # Summary data
    total_carbon_tonnes: float
    total_co2e_tonnes: float
    uncertainty_pct: float
    area_hectares: float
    tree_count: int

    # By pool
    pools: dict[str, dict[str, float]]

    # Credits
    credits: dict[str, Any] | None

    # Audit
    audit_id: str
    methodology_version: str

    # Files
    pdf_path: str | None = None
    excel_path: str | None = None

    # Metadata
    generated_at: datetime = field(default_factory=datetime.utcnow)
    config: CarbonReportConfig = field(default_factory=CarbonReportConfig)


class CarbonReportGenerator:
    """
    Generates carbon stock reports in PDF and Excel formats.

    Supports VCS, CAR, and ACR protocol-compliant reporting.
    """

    def __init__(self, output_directory: str | None = None):
        """
        Initialize report generator.

        Args:
            output_directory: Directory for output files
        """
        self.output_directory = Path(output_directory) if output_directory else Path("./reports")
        self.output_directory.mkdir(parents=True, exist_ok=True)

    def generate_report(
        self,
        carbon_stock: ProjectCarbonStock,
        tree_estimates: list[TreeCarbonEstimate] | None = None,
        credits: dict[str, Any] | None = None,
        config: CarbonReportConfig | None = None,
        output_format: str = "both",  # 'pdf', 'excel', 'both'
    ) -> CarbonReport:
        """
        Generate a carbon stock report.

        Args:
            carbon_stock: Project carbon stock data
            tree_estimates: Individual tree estimates (optional)
            credits: Carbon credit calculations (optional)
            config: Report configuration
            output_format: Output format(s)

        Returns:
            CarbonReport with generated files
        """
        if config is None:
            config = CarbonReportConfig()

        report_id = f"CARBON-RPT-{datetime.utcnow().strftime('%Y%m%d-%H%M%S')}"

        logger.info(f"Generating carbon report {report_id} for {carbon_stock.project_id}")

        # Prepare pool data
        pools_data = {}
        for pool_type, pool in carbon_stock.pools.items():
            pools_data[pool_type.value] = {
                "carbon_tonnes": pool.carbon_tonnes.value,
                "co2e_tonnes": pool.co2e_tonnes.value,
                "uncertainty_pct": pool.carbon_tonnes.uncertainty_pct,
                "carbon_density_t_ha": pool.carbon_density_t_ha,
            }

        # Create report object
        report = CarbonReport(
            report_id=report_id,
            project_id=carbon_stock.project_id,
            analysis_id=carbon_stock.analysis_id,
            protocol=carbon_stock.protocol,
            total_carbon_tonnes=carbon_stock.total_carbon_tonnes.value,
            total_co2e_tonnes=carbon_stock.total_co2e_tonnes.value,
            uncertainty_pct=carbon_stock.total_carbon_tonnes.uncertainty_pct,
            area_hectares=carbon_stock.area_hectares,
            tree_count=carbon_stock.tree_count,
            pools=pools_data,
            credits=credits,
            audit_id=carbon_stock.audit_id,
            methodology_version=carbon_stock.methodology_version,
            config=config,
        )

        # Generate PDF
        if output_format in ("pdf", "both"):
            try:
                pdf_path = self._generate_pdf(report, tree_estimates)
                report.pdf_path = str(pdf_path)
            except Exception as e:
                logger.warning(f"Failed to generate PDF: {e}")

        # Generate Excel
        if output_format in ("excel", "both"):
            try:
                excel_path = self._generate_excel(report, tree_estimates)
                report.excel_path = str(excel_path)
            except Exception as e:
                logger.warning(f"Failed to generate Excel: {e}")

        logger.info(f"Generated carbon report: {report_id}")
        return report

    def _generate_pdf(
        self,
        report: CarbonReport,
        tree_estimates: list[TreeCarbonEstimate] | None = None,
    ) -> Path:
        """Generate PDF report."""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter
            from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
            from reportlab.lib.units import inch
            from reportlab.platypus import (
                Paragraph,
                SimpleDocTemplate,
                Spacer,
                Table,
                TableStyle,
            )
        except ImportError:
            logger.warning("reportlab not installed, skipping PDF generation")
            raise ImportError("reportlab required for PDF generation")

        pdf_path = self.output_directory / f"{report.report_id}.pdf"

        doc = SimpleDocTemplate(
            str(pdf_path),
            pagesize=letter,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=72,
        )

        styles = getSampleStyleSheet()
        story = []

        # Title
        title_style = ParagraphStyle(
            "Title",
            parent=styles["Heading1"],
            fontSize=18,
            spaceAfter=20,
        )
        story.append(Paragraph("Carbon Stock Assessment Report", title_style))
        story.append(Spacer(1, 12))

        # Protocol badge
        protocol_name = {
            CarbonProtocol.VCS: "Verified Carbon Standard (VCS/Verra)",
            CarbonProtocol.CAR: "Climate Action Reserve (CAR)",
            CarbonProtocol.ACR: "American Carbon Registry (ACR)",
            CarbonProtocol.FIA: "USFS Forest Inventory and Analysis",
        }
        story.append(Paragraph(
            f"<b>Protocol:</b> {protocol_name.get(report.protocol, report.protocol.value)}",
            styles["Normal"]
        ))
        story.append(Paragraph(
            f"<b>Methodology:</b> {report.methodology_version}",
            styles["Normal"]
        ))
        story.append(Paragraph(
            f"<b>Report ID:</b> {report.report_id}",
            styles["Normal"]
        ))
        story.append(Paragraph(
            f"<b>Audit ID:</b> {report.audit_id}",
            styles["Normal"]
        ))
        story.append(Spacer(1, 20))

        # Summary section
        story.append(Paragraph("Executive Summary", styles["Heading2"]))
        story.append(Spacer(1, 10))

        summary_data = [
            ["Metric", "Value", "Unit"],
            ["Project ID", report.project_id, ""],
            ["Analysis ID", report.analysis_id, ""],
            ["Total Area", f"{report.area_hectares:.2f}", "hectares"],
            ["Tree Count", f"{report.tree_count:,}", "trees"],
            ["Total Carbon Stock", f"{report.total_carbon_tonnes:.2f}", "tonnes C"],
            ["Total CO₂ Equivalent", f"{report.total_co2e_tonnes:.2f}", "tonnes CO₂e"],
            ["Uncertainty (95% CI)", f"±{report.uncertainty_pct:.1f}", "%"],
        ]

        summary_table = Table(summary_data, colWidths=[2.5*inch, 2*inch, 1.5*inch])
        summary_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E7D32")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#E8F5E9")),
            ("GRID", (0, 0), (-1, -1), 1, colors.HexColor("#81C784")),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 20))

        # Carbon pools section
        story.append(Paragraph("Carbon Pools", styles["Heading2"]))
        story.append(Spacer(1, 10))

        pool_names = {
            "above_ground_live": "Above-Ground Live Biomass",
            "below_ground_live": "Below-Ground Live Biomass",
            "dead_wood": "Dead Wood",
            "litter": "Litter",
            "soil": "Soil Organic Carbon",
        }

        pool_data = [["Pool", "Carbon (t)", "CO₂e (t)", "Density (tC/ha)", "Uncertainty"]]
        for pool_id, data in report.pools.items():
            pool_data.append([
                pool_names.get(pool_id, pool_id),
                f"{data['carbon_tonnes']:.2f}",
                f"{data['co2e_tonnes']:.2f}",
                f"{data['carbon_density_t_ha']:.2f}",
                f"±{data['uncertainty_pct']:.1f}%",
            ])

        pool_table = Table(pool_data, colWidths=[2.5*inch, 1*inch, 1*inch, 1.2*inch, 1*inch])
        pool_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1565C0")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#E3F2FD")),
            ("GRID", (0, 0), (-1, -1), 1, colors.HexColor("#64B5F6")),
        ]))
        story.append(pool_table)
        story.append(Spacer(1, 20))

        # Carbon credits section (if included)
        if report.credits and report.config.include_credits:
            story.append(Paragraph("Carbon Credits Potential", styles["Heading2"]))
            story.append(Spacer(1, 10))

            credits_data = [
                ["Metric", "Value"],
                ["Gross CO₂e", f"{report.credits['gross_co2e_tonnes']:.2f} tonnes"],
                ["Conservative Deduction", f"{report.credits['conservative_deduction_pct']:.0f}%"],
                ["Net CO₂e", f"{report.credits['net_co2e_tonnes']:.2f} tonnes"],
                ["Potential Credits", f"{report.credits['credits']:.2f}"],
                ["Registry", report.credits['registry'].upper()],
            ]

            if "estimated_value_usd" in report.credits:
                values = report.credits["estimated_value_usd"]
                credits_data.append([
                    "Estimated Value (USD)",
                    f"${values['low']:,.0f} - ${values['high']:,.0f}",
                ])

            credits_table = Table(credits_data, colWidths=[3*inch, 3*inch])
            credits_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FF8F00")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#FFF3E0")),
                ("GRID", (0, 0), (-1, -1), 1, colors.HexColor("#FFB74D")),
            ]))
            story.append(credits_table)
            story.append(Spacer(1, 20))

        # Methodology section
        if report.config.include_methodology:
            story.append(Paragraph("Methodology", styles["Heading2"]))
            story.append(Spacer(1, 10))

            methodology_text = f"""
            This carbon stock assessment was conducted following the {protocol_name.get(report.protocol, '')}
            methodology ({report.methodology_version}).

            <b>Above-Ground Biomass:</b> Calculated using Jenkins et al. (2003) national-scale allometric
            equations based on diameter at breast height (DBH).

            <b>Below-Ground Biomass:</b> Estimated using IPCC (2006) default root-to-shoot ratios
            for temperate forests (R:S = 0.26).

            <b>Carbon Conversion:</b> Biomass converted to carbon using IPCC default carbon fraction
            of {report.protocol == CarbonProtocol.CAR and '0.50' or '0.47'}.

            <b>CO₂ Equivalent:</b> Carbon converted to CO₂ equivalent using molecular weight ratio (44/12 = 3.667).

            <b>Uncertainty:</b> Quantified using IPCC Tier 1 error propagation methodology, combining
            measurement uncertainty (±5% DBH) with equation uncertainty (±30% biomass equations).
            """
            story.append(Paragraph(methodology_text, styles["Normal"]))
            story.append(Spacer(1, 20))

        # Footer
        story.append(Spacer(1, 30))
        footer_style = ParagraphStyle(
            "Footer",
            parent=styles["Normal"],
            fontSize=8,
            textColor=colors.gray,
        )
        story.append(Paragraph(
            f"Generated: {report.generated_at.strftime('%Y-%m-%d %H:%M:%S UTC')}",
            footer_style
        ))
        story.append(Paragraph(
            "LiDAR Forest Analysis Platform - Carbon Stock Assessment Module",
            footer_style
        ))

        doc.build(story)
        logger.info(f"Generated PDF report: {pdf_path}")
        return pdf_path

    def _generate_excel(
        self,
        report: CarbonReport,
        tree_estimates: list[TreeCarbonEstimate] | None = None,
    ) -> Path:
        """Generate Excel report."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.warning("openpyxl not installed, skipping Excel generation")
            raise ImportError("openpyxl required for Excel generation")

        excel_path = self.output_directory / f"{report.report_id}.xlsx"

        wb = Workbook()

        # Styles
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # === Summary Sheet ===
        ws = wb.active
        ws.title = "Summary"

        # Title
        ws["A1"] = "Carbon Stock Assessment Report"
        ws["A1"].font = Font(bold=True, size=16)
        ws.merge_cells("A1:D1")

        # Protocol info
        ws["A3"] = "Protocol:"
        ws["B3"] = report.protocol.value.upper()
        ws["A4"] = "Methodology:"
        ws["B4"] = report.methodology_version
        ws["A5"] = "Report ID:"
        ws["B5"] = report.report_id
        ws["A6"] = "Audit ID:"
        ws["B6"] = report.audit_id

        # Summary table
        summary_start = 8
        summary_headers = ["Metric", "Value", "Unit"]
        summary_data = [
            ["Project ID", report.project_id, ""],
            ["Analysis ID", report.analysis_id, ""],
            ["Total Area", f"{report.area_hectares:.2f}", "hectares"],
            ["Tree Count", f"{report.tree_count:,}", "trees"],
            ["Total Carbon Stock", f"{report.total_carbon_tonnes:.2f}", "tonnes C"],
            ["Total CO₂ Equivalent", f"{report.total_co2e_tonnes:.2f}", "tonnes CO₂e"],
            ["Uncertainty (95% CI)", f"±{report.uncertainty_pct:.1f}", "%"],
        ]

        for col, header in enumerate(summary_headers, 1):
            cell = ws.cell(row=summary_start, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border

        for row_idx, row_data in enumerate(summary_data, summary_start + 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

        # Set column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 15

        # === Carbon Pools Sheet ===
        ws_pools = wb.create_sheet("Carbon Pools")

        pool_headers = ["Pool", "Carbon (t)", "CO₂e (t)", "Density (tC/ha)", "Uncertainty (%)"]
        for col, header in enumerate(pool_headers, 1):
            cell = ws_pools.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border

        pool_names = {
            "above_ground_live": "Above-Ground Live Biomass",
            "below_ground_live": "Below-Ground Live Biomass",
            "dead_wood": "Dead Wood",
            "litter": "Litter",
            "soil": "Soil Organic Carbon",
        }

        for row_idx, (pool_id, data) in enumerate(report.pools.items(), 2):
            ws_pools.cell(row=row_idx, column=1, value=pool_names.get(pool_id, pool_id)).border = thin_border
            ws_pools.cell(row=row_idx, column=2, value=data["carbon_tonnes"]).border = thin_border
            ws_pools.cell(row=row_idx, column=3, value=data["co2e_tonnes"]).border = thin_border
            ws_pools.cell(row=row_idx, column=4, value=data["carbon_density_t_ha"]).border = thin_border
            ws_pools.cell(row=row_idx, column=5, value=data["uncertainty_pct"]).border = thin_border

        for col in range(1, 6):
            ws_pools.column_dimensions[get_column_letter(col)].width = 25

        # === Credits Sheet (if included) ===
        if report.credits and report.config.include_credits:
            ws_credits = wb.create_sheet("Carbon Credits")

            credits_headers = ["Metric", "Value"]
            for col, header in enumerate(credits_headers, 1):
                cell = ws_credits.cell(row=1, column=col, value=header)
                cell.fill = PatternFill(start_color="FF8F00", end_color="FF8F00", fill_type="solid")
                cell.font = header_font
                cell.border = thin_border

            credits_rows = [
                ["Gross CO₂e (tonnes)", report.credits["gross_co2e_tonnes"]],
                ["Conservative Deduction (%)", report.credits["conservative_deduction_pct"]],
                ["Net CO₂e (tonnes)", report.credits["net_co2e_tonnes"]],
                ["Potential Credits", report.credits["credits"]],
                ["Registry", report.credits["registry"].upper()],
                ["Methodology", report.credits["methodology"]],
            ]

            if "estimated_value_usd" in report.credits:
                credits_rows.append([
                    "Estimated Value Low (USD)",
                    report.credits["estimated_value_usd"]["low"],
                ])
                credits_rows.append([
                    "Estimated Value Mid (USD)",
                    report.credits["estimated_value_usd"]["mid"],
                ])
                credits_rows.append([
                    "Estimated Value High (USD)",
                    report.credits["estimated_value_usd"]["high"],
                ])

            for row_idx, row_data in enumerate(credits_rows, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws_credits.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border

            ws_credits.column_dimensions["A"].width = 30
            ws_credits.column_dimensions["B"].width = 20

        # === Tree-Level Sheet (if included and available) ===
        if report.config.include_tree_level and tree_estimates:
            ws_trees = wb.create_sheet("Tree Carbon")

            tree_headers = [
                "Tree ID", "Species", "DBH (cm)", "Height (m)",
                "AGB (kg)", "BGB (kg)", "Carbon (kg)", "CO₂e (kg)", "Uncertainty (%)"
            ]
            for col, header in enumerate(tree_headers, 1):
                cell = ws_trees.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border

            for row_idx, tree in enumerate(tree_estimates[:10000], 2):  # Limit to 10000 trees
                ws_trees.cell(row=row_idx, column=1, value=tree.tree_id).border = thin_border
                ws_trees.cell(row=row_idx, column=2, value=tree.species_code or "").border = thin_border
                ws_trees.cell(row=row_idx, column=3, value=tree.dbh_cm).border = thin_border
                ws_trees.cell(row=row_idx, column=4, value=tree.height_m).border = thin_border
                ws_trees.cell(row=row_idx, column=5, value=tree.aboveground_biomass_kg.value).border = thin_border
                ws_trees.cell(row=row_idx, column=6, value=tree.belowground_biomass_kg.value).border = thin_border
                ws_trees.cell(row=row_idx, column=7, value=tree.carbon_kg.value).border = thin_border
                ws_trees.cell(row=row_idx, column=8, value=tree.co2e_kg.value).border = thin_border
                ws_trees.cell(row=row_idx, column=9, value=tree.carbon_kg.uncertainty_pct).border = thin_border

            for col in range(1, 10):
                ws_trees.column_dimensions[get_column_letter(col)].width = 15

        # === Audit Trail Sheet ===
        if report.config.include_audit_trail:
            ws_audit = wb.create_sheet("Audit Trail")

            ws_audit["A1"] = "Audit ID"
            ws_audit["B1"] = report.audit_id
            ws_audit["A2"] = "Report ID"
            ws_audit["B2"] = report.report_id
            ws_audit["A3"] = "Project ID"
            ws_audit["B3"] = report.project_id
            ws_audit["A4"] = "Analysis ID"
            ws_audit["B4"] = report.analysis_id
            ws_audit["A5"] = "Protocol"
            ws_audit["B5"] = report.protocol.value.upper()
            ws_audit["A6"] = "Methodology"
            ws_audit["B6"] = report.methodology_version
            ws_audit["A7"] = "Generated At"
            ws_audit["B7"] = report.generated_at.isoformat()

            for row in range(1, 8):
                ws_audit[f"A{row}"].font = Font(bold=True)

            ws_audit.column_dimensions["A"].width = 20
            ws_audit.column_dimensions["B"].width = 40

        wb.save(excel_path)
        logger.info(f"Generated Excel report: {excel_path}")
        return excel_path

    def generate_summary_dict(self, report: CarbonReport) -> dict[str, Any]:
        """
        Generate a summary dictionary for API responses.

        Args:
            report: Carbon report

        Returns:
            Dictionary with report summary
        """
        return {
            "report_id": report.report_id,
            "project_id": report.project_id,
            "analysis_id": report.analysis_id,
            "protocol": report.protocol.value,
            "methodology_version": report.methodology_version,
            "audit_id": report.audit_id,
            "summary": {
                "total_carbon_tonnes": report.total_carbon_tonnes,
                "total_co2e_tonnes": report.total_co2e_tonnes,
                "uncertainty_pct": report.uncertainty_pct,
                "area_hectares": report.area_hectares,
                "tree_count": report.tree_count,
            },
            "pools": report.pools,
            "credits": report.credits,
            "files": {
                "pdf": report.pdf_path,
                "excel": report.excel_path,
            },
            "generated_at": report.generated_at.isoformat(),
        }
